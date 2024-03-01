from django.http import HttpResponse, Http404
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from .models import Quiz, QuizTaker, Result, Answer
from django.http import HttpResponse
from django.shortcuts import render, redirect
from .models import *
from django.contrib.auth import login, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from django.http import JsonResponse
from .models import Quiz, Question, Option, Answer

def chart_data_view(request):
    # Assuming you have a model for storing quiz data, adjust this query accordingly
    questions = Question.objects.all()

    # Calculate the number of true solvers for each question
    data = []
    for question in questions:
        true_solvers = Answer.objects.filter(option__question=question, option__is_correct=True).count()
        data.append({
            'question_id': question.id,
            'true_solvers': true_solvers
        })

    return JsonResponse(data, safe=False)




@login_required(login_url = 'dash:login')
def main(request):
    quizes = Quiz.objects.filter(author = request.user)
    context = {
        "quizes" : quizes
    }
    return render(request, 'main.html', context)

# creators
@login_required(login_url = 'dash:login')
def create_quiz(request):
    if request.method == 'POST':
        title = request.POST['title']
        quiz = Quiz.objects.create(
            title = title,
            author = request.user
        )
        return redirect('dash:quest_create', quiz.id)
    return render(request, 'quiz/create-quiz.html')

@login_required(login_url = 'dash:login')
def add_question(request, id):
    quiz = Quiz.objects.get(id = id)
    if request.method == 'POST':
        title = request.POST['title']
        question = Question.objects.create(
            quiz = quiz,
            title = title
        )
        correct_answer = request.POST['correct']
        Option.objects.create(
            question = question,
            name = correct_answer,
            is_correct = True
        )
        incorrect_answers = [request.POST[key] for key in request.POST if key.startswith('incorrect')]
        for incorrect_answer in incorrect_answers:
            Option.objects.create(
                question = question,
                name = incorrect_answer,
                is_correct = False
            )
        if request.POST['submit_action'] == 'exit':
            return redirect('dash:main')
    return render(request, 'details/add-question.html')

@login_required(login_url='dash:login')
def create_question(request, id):
    quiz = Quiz.objects.get(id=id)
    if request.method == 'POST':
        title = request.POST['title']
        ques = Question.objects.create(
            quiz=quiz,
            title=title
        )
        correct_answer = request.POST['correct']
        Option.objects.create(
            question=ques,
            name=correct_answer,
            is_correct=True
        )
        incorrect_answers = [request.POST[key] for key in request.POST if key.startswith('incorrect')]
        
        for incorrect_answer in incorrect_answers:
            Option.objects.create(
                question=ques,
                name=incorrect_answer,
                is_correct=False
            )
        if request.POST['submit_action'] == 'exit':
            return redirect('dash:main')
    return render(request, 'quiz/create-question.html')


# question list
@login_required(login_url = 'dash:login')
def questions_list(request, id):
    correct_answer = Option.objects.filter(question__quiz_id = id, is_correct = True)
    quests = Question.objects.filter(quiz_id = id)
    context = {
        'questions': quests,
        'correct_answer': correct_answer[0]
    }
    return render (request, 'details/questions.html', context)

@login_required(login_url='dash:login')
def quest_detail(request, id):
    question = Question.objects.get(id=id)
    option_correct = Option.objects.get(question=question, is_correct=True)
    options = Option.objects.filter(question=question, is_correct=False).order_by('id')
    context = {
        'question': question,
        'options': options,
        'option_correct': option_correct
    }
    if request.method == 'POST':
        question.title = request.POST['title']
        question.save()

        option_correct.name = request.POST['correct']
        option_correct.save()

        # Extract all incorrect answers dynamically
        incorrect_answers = []
        for key, value in request.POST.items():
            if key.startswith('incorrect'):
                incorrect_answers.append(value)

        for i, opt in enumerate(options):
            if i < len(incorrect_answers):
                opt.name = incorrect_answers[i]
                opt.save()

    return render(request, 'details/detail.html', context)




@login_required(login_url = 'dash:login')
def quiz_delete(request, id):
    Quiz.objects.get(id = id).delete()
    return redirect('dash:main')

@login_required(login_url = 'dash:login')
def get_results(request, id):
    quiz = Quiz.objects.get(id=id)
    taker = QuizTaker.objects.filter(quiz=quiz)
    results = []
    for i in taker:
        result = Result.objects.filter(taker=i)
        if result.exists():
            results.append(result.first())
    context = {
        'results': results,
        'quiz': quiz
    }
    return render(request, 'quiz/results.html', context)


def result_detail(request, id):
    result = Result.objects.get(id=id)
    answers = Answer.objects.filter(taker=result.taker)
    context = {
        'taker':result.taker,
        'answers':answers
    }
    return render(request, 'quiz/result-detail.html', context)


# excel downloader

from operator import attrgetter

def get_results_excel_downloader(request, id):
    try:
        quiz = Quiz.objects.get(id=id)
    except Quiz.DoesNotExist:
        raise Http404("Quiz does not exist")

    takers = QuizTaker.objects.filter(quiz=quiz)
    results = []
    for taker in takers:
        result = Result.objects.filter(taker=taker)
        if result.exists():
            results.append(result.first())

    # Sort results by percentage
    results.sort(key=lambda result: result.percentage, reverse=True)

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    headers = ['Name', 'Phone', 'Questions', 'Correct', 'Incorrect', '%']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = header
        sheet[f'{col_letter}1'].alignment = Alignment(horizontal='center')

    for row_num, result in enumerate(results, 2):
        taker = result.taker
        answers = Answer.objects.filter(taker=taker)
        total_questions = answers.count()
        correct_answers = answers.filter(is_correct=True).count()
        incorrect_answers = total_questions - correct_answers

        sheet[f'A{row_num}'] = taker.full_name
        sheet[f'B{row_num}'] = taker.phone
        sheet[f'C{row_num}'] = total_questions
        sheet[f'D{row_num}'] = correct_answers
        sheet[f'E{row_num}'] = incorrect_answers
        sheet[f'F{row_num}'] = result.percentage

    # Save the workbook as an Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=results.xlsx'
    workbook.save(response)

    return response


#login
def logging_in(request):
    status = False
    if request.method == 'POST':
        username = request.POST['username']
        password =  request.POST['password']
        user = authenticate(username = username, password=password)
        if user:
            login(request, user)
            return redirect('dash:main')
        else:
            status = 'incorrect username or password'
    return render(request, 'auth/login.html', {'status':status})

def register(request):
    status = False
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        if not User.objects.filter(username = username).first():
            User.objects.create_user(
                username=username,
                password=password
            )
            user = authenticate(username = username, password = password)
            login(request, user)
            return redirect('dash:main')
        else:
            status  = f'the username {username} is occupied'
    return render(request, 'auth/register.html', {'status': status})

